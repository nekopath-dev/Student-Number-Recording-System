import nfc
import re
import pygame
import time
from openpyxl import load_workbook
from datetime import datetime

pygame.mixer.init()

current_step = "customer"

print("個別ブース利用登録を行います。利用者の方は、ICカードリーダーに学生証をかざしてください。")

def play_sound(file_path):
    pygame.mixer.music.load(file_path)
    pygame.mixer.music.play()

file_path = r"path/to/excel/file.xlsx" # 記録するExcelファイルパスを入力
wb = load_workbook(file_path)
ws = wb.active

def on_connect(tag: nfc.tag.Tag) -> bool:
    global current_step
    print("カード情報を読み取っています。かざしたままお待ち下さい。")
    data = "\n".join(tag.dump())
    pattern = re.compile(r'[A-Z]{2}\d{5}') # 学生番号の正規表現パターンを変更
    match = pattern.search(data)
    if match:
        student_id = match.group()
        print("Student ID:", student_id)

        now = datetime.now()
        date = now.strftime('%Y-%m-%d')
        current_time = now.strftime('%H:%M')  # 変数名を `time` から `current_time` に変更

        if current_step == "customer":
            ws.insert_rows(3)
            ws['A3'] = date
            ws['B3'] = current_time  # 更新された変数名を使用
            ws['C3'] = student_id
            print("利用者登録が完了しました。続けて、担当者の学生証をかざしてください")
            play_sound(r"path/to/sound/file.mp3") # 利用者登録完了時の音声ファイルパスを入力
            current_step = "manager"
            return True
        elif current_step == "manager":
            ws['E3'] = student_id
            wb.save(file_path)
            play_sound(r"path/to/sound/file.mp3") # 担当者登録完了時の音声ファイルパスを入力
            print("処理中です。しばらくお待ち下さい。")
            time.sleep(1)  # 1秒間待つ
            print("担当者登録が完了しました。")
            raise SystemExit("登録作業は完了です。プログラムを終了します。")
    else:
        print("学生番号の読み取りに失敗しました。学生証かどうかを再度確認してからやり直してください。")
        return True  # 学生番号が見つからない場合は、再度読み取りを続ける

def on_release(tag: nfc.tag.Tag) -> None:
    print("カードが離されました。")

# Main NFC polling loop
with nfc.ContactlessFrontend('usb') as clf:
    try:
        while True:
            clf.connect(rdwr={"on-connect": on_connect, "on-release": on_release})
    except SystemExit as e:
        print(e)
